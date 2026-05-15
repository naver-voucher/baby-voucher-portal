"""
COS 메일 발송 로컬 프록시 서버.
브라우저(index.html) → localhost:5000/send (JSON) → COS API (multipart)
"""
import base64
import csv
import io
import json
import os
import platform
import tempfile
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from urllib.parse import urlparse

import requests

COS_API_URL = "https://cos.navercorp.com/cos/mail/v2/api"
PORT = 5001
BASE_DIR = Path(__file__).parent
HISTORY_FILE = BASE_DIR / "history.json"


class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        print(f"[proxy] {self.address_string()} - {fmt % args}")

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        if path in ("/", "/index.html"):
            self._serve_file(BASE_DIR / "index.html", "text/html; charset=utf-8")
        elif path == "/template":
            self._serve_template()
        elif path == "/history":
            self._handle_get_history()
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/send":
            self._handle_send()
        elif path == "/upload-list":
            self._handle_upload_list()
        elif path == "/history":
            self._handle_save_history()
        else:
            self.send_response(404)
            self.end_headers()

    # ── POST /send ──────────────────────────────────────────────
    def _handle_send(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length)

        try:
            body = json.loads(raw)
        except Exception:
            self._json_response(400, {"state": "ERROR", "errorMessage": "요청 파싱 실패"})
            return

        mail_request = body.get("mailRequest", {})
        attach = body.get("attachFile")  # {"name":"..","data":"base64.."} or {"name":"..","path":".."}

        mail_json = json.dumps(mail_request, ensure_ascii=False)
        files = {"mailRequest": (None, mail_json, "application/json")}

        tmp_path = None
        attach_fp = None
        try:
            if attach:
                if attach.get("path"):
                    local_path = Path(attach["path"])
                    if not local_path.is_file():
                        self._json_response(400, {"state": "ERROR", "errorMessage": f"첨부 파일 없음: {local_path}"})
                        return
                    attach_fp = local_path.open("rb")
                    files["attachFile"] = (local_path.name, attach_fp)
                elif attach.get("data"):
                    file_bytes = base64.b64decode(attach["data"])
                    suffix = Path(attach.get("name", "file")).suffix
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
                    tmp.write(file_bytes)
                    tmp.close()
                    tmp_path = tmp.name
                    files["attachFile"] = (attach.get("name", "file"), open(tmp_path, "rb"))

            resp = requests.post(COS_API_URL, files=files, timeout=30)
            print(f"[proxy] COS 응답: {resp.status_code} {resp.text[:200]}")
            self._json_response(resp.status_code, resp.json() if resp.content else {})

        except requests.exceptions.ConnectionError as e:
            self._json_response(502, {"state": "ERROR", "errorMessage": f"COS 서버 연결 실패: {e}"})
        except Exception as e:
            self._json_response(500, {"state": "ERROR", "errorMessage": str(e)})
        finally:
            if attach_fp:
                try:
                    attach_fp.close()
                except Exception:
                    pass
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    # ── POST /upload-list ───────────────────────────────────────
    def _handle_upload_list(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length)

        try:
            body = json.loads(raw)
        except Exception:
            self._json_response(400, {"state": "ERROR", "errorMessage": "요청 파싱 실패"})
            return

        try:
            file_bytes = base64.b64decode(body.get("data", ""))
        except Exception:
            self._json_response(400, {"state": "ERROR", "errorMessage": "파일 디코딩 실패"})
            return

        file_name = body.get("name", "list.xlsx")
        ext = Path(file_name).suffix.lower()
        rows = []

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        continue  # skip header
                    if row and row[0]:
                        rows.append({
                            "email": str(row[0]).strip(),
                            "path":  str(row[1]).strip() if len(row) > 1 and row[1] else "",
                        })
            except ImportError:
                self._json_response(500, {"state": "ERROR", "errorMessage": "openpyxl 미설치 — pip install openpyxl 후 재시작하거나 CSV 파일을 사용하세요"})
                return
        else:
            text = file_bytes.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            for i, row in enumerate(reader):
                if i == 0:
                    continue  # skip header
                if row and row[0]:
                    rows.append({
                        "email": row[0].strip(),
                        "path":  row[1].strip() if len(row) > 1 else "",
                    })

        self._json_response(200, {"rows": rows})

    # ── GET /template ───────────────────────────────────────────
    def _serve_template(self):
        is_mac = platform.system() == "Darwin"
        example_path = "/Users/username/Downloads/report.zip" if is_mac else r"C:\Users\USER\Downloads\report.zip"
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "수신자목록"
            ws.append(["수신자이메일", "첨부파일경로"])
            ws.append(["partner@example.com", example_path])
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 55
            buf = io.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",
                "attachment; filename*=UTF-8''%EC%88%98%EC%8B%A0%EC%9E%90%EB%AA%A9%EB%A1%9D.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except ImportError:
            header = "수신자이메일,첨부파일경로\n"
            example = f"partner@example.com,{example_path}\n"
            data = (header + example).encode("utf-8-sig")
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition",
                "attachment; filename*=UTF-8''%EC%88%98%EC%8B%A0%EC%9E%90%EB%AA%A9%EB%A1%9D.csv")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

    # ── GET /history ────────────────────────────────────────────
    def _handle_get_history(self):
        data = HISTORY_FILE.read_bytes() if HISTORY_FILE.is_file() else b"[]"
        self.send_response(200)
        self._cors()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    # ── POST /history ────────────────────────────────────────────
    def _handle_save_history(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length)
        try:
            history = json.loads(raw)
            HISTORY_FILE.write_text(
                json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._json_response(200, {"ok": True})
        except Exception as e:
            self._json_response(500, {"ok": False, "error": str(e)})

    # ── helpers ─────────────────────────────────────────────────
    def _json_response(self, status: int, data: dict):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self._cors()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _serve_file(self, path: Path, content_type: str):
        if not path.is_file():
            self.send_response(404)
            self.end_headers()
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


if __name__ == "__main__":
    server = HTTPServer(("localhost", PORT), Handler)
    print(f"[proxy] 서버 시작: http://localhost:{PORT}")
    print("[proxy] 종료하려면 Ctrl+C")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[proxy] 종료")
